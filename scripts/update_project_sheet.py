from datetime import datetime

from openpyxl import load_workbook


PATH = "docs/Group 1 - SWE Project Sheet.xlsx"


wb = load_workbook(PATH)


def set_row(ws, row, values):
    for col, value in values.items():
        ws[f"{col}{row}"] = value


# SRS / workbook update target date
TODAY = datetime(2026, 4, 3)


def dt(year, month, day):
    return datetime(year, month, day)


# ---------------------------------------------------------------------------
# Implementation
# ---------------------------------------------------------------------------
ws = wb["Implementation"]

if ws["I4"].value != "Commit 9e3e7c0 (2026-03-25)":
    ws.insert_rows(4, 8)

set_row(
    ws,
    3,
    {
        "A": "NFR-01, NFR-02, NFR-03",
        "B": "QA & Test Infrastructure",
        "C": "-",
        "D": "smoke_test.sh (59 smoke checks), test_cases.csv (48 documented TCs), e2e-full-flow.js, VHS tape recordings",
        "E": dt(2026, 3, 18),
        "F": dt(2026, 3, 25),
        "G": dt(2026, 3, 25),
        "H": "Expanded automated test coverage to 59 smoke checks, retained 48 categorized documented CSV test cases, and added full-system end-to-end flow validation",
        "I": "Commits 8f0f443, f2b7f20, ec4a4cb, 3fac9c2",
        "J": "Lab 7 & 8 Testing Assignment + Sprint 5 Integration",
        "K": "Nayan, Chaitanya, Yash",
        "L": "Chaitanya, Nayan",
        "M": "Complete",
        "N": "#16",
        "O": dt(2026, 3, 18),
        "P": dt(2026, 3, 25),
    },
)

for cell in ("H18",):
    ws[cell] = None

implementation_rows = [
    (
        4,
        {
            "A": "NFR-Security",
            "B": "Auth Manager",
            "C": "AuthService, LDAP Docker Setup",
            "D": "loginUser, LDAP bind validation",
            "E": dt(2026, 3, 25),
            "F": dt(2026, 3, 25),
            "G": dt(2026, 3, 25),
            "H": "Replaced the LDAP fallback-only behavior with local Dockerized LDAP-backed authentication for seeded users",
            "I": "Commit 9e3e7c0 (2026-03-25)",
            "J": "Sprint 5 Security / Integration",
            "K": "Yash, Chaitanya",
            "L": "Yash, Nayan",
            "M": "Complete",
            "N": "#17",
            "O": dt(2026, 3, 25),
            "P": dt(2026, 3, 25),
        },
    ),
    (
        5,
        {
            "A": "NFR-Integration, NFR-Documentation",
            "B": "QA & Integration",
            "C": "E2E Runner, API Documentation",
            "D": "test:e2e, migrate:deploy, route index maintenance",
            "E": dt(2026, 3, 25),
            "F": dt(2026, 3, 25),
            "G": dt(2026, 3, 25),
            "H": "Added full-system end-to-end flow runner and API documentation for backend integration closeout",
            "I": "Commits 3fac9c2, 2d7e421 (2026-03-25)",
            "J": "Sprint 5 Integration",
            "K": "Yash, Chaitanya",
            "L": "Yash, Nayan",
            "M": "Complete",
            "N": "#18",
            "O": dt(2026, 3, 25),
            "P": dt(2026, 3, 25),
        },
    ),
    (
        6,
        {
            "A": "REQ-40, REQ-41",
            "B": "Notification & Availability",
            "C": "DoctorController, NotificationController",
            "D": "markUnavailable, listMyNotifications, markNotificationRead",
            "E": dt(2026, 3, 25),
            "F": dt(2026, 3, 25),
            "G": dt(2026, 3, 25),
            "H": "Added physician absence workflow and in-app availability notifications for affected users",
            "I": "Commit 9182fcc (2026-03-25)",
            "J": "Sprint 5 Implementation",
            "K": "Yash, Nayan",
            "L": "Chaitanya, Nayan",
            "M": "Complete",
            "N": "#19",
            "O": dt(2026, 3, 25),
            "P": dt(2026, 3, 25),
        },
    ),
    (
        7,
        {
            "A": "REQ-49",
            "B": "Admin User Management",
            "C": "AdminController",
            "D": "createUser, listUsers, updateUser",
            "E": dt(2026, 3, 24),
            "F": dt(2026, 3, 24),
            "G": dt(2026, 3, 24),
            "H": "Added admin user creation, filtering, role update, deactivation, and account status controls",
            "I": "Commit 5ff16c2 (2026-03-24)",
            "J": "Sprint 5 Implementation",
            "K": "Yash, Chaitanya",
            "L": "Chaitanya, Nayan",
            "M": "Complete",
            "N": "#20",
            "O": dt(2026, 3, 24),
            "P": dt(2026, 3, 24),
        },
    ),
    (
        8,
        {
            "A": "REQ-50, REQ-52",
            "B": "Reporting & Admin",
            "C": "ReportController",
            "D": "usage, attendanceSummary",
            "E": dt(2026, 3, 24),
            "F": dt(2026, 3, 24),
            "G": dt(2026, 3, 24),
            "H": "Added usage and attendance summary reporting endpoints for admin review",
            "I": "Commit 248d70a (2026-03-24)",
            "J": "Sprint 5 Implementation",
            "K": "Yash, Nayan",
            "L": "Chaitanya, Nayan",
            "M": "Complete",
            "N": "#21",
            "O": dt(2026, 3, 24),
            "P": dt(2026, 3, 24),
        },
    ),
    (
        9,
        {
            "A": "REQ-51",
            "B": "Public Information & Admin",
            "C": "EventController",
            "D": "publish, listUpcomingEvents",
            "E": dt(2026, 3, 24),
            "F": dt(2026, 3, 24),
            "G": dt(2026, 3, 24),
            "H": "Added PHC event publishing and public upcoming-events listing",
            "I": "Commit b66f7f2 (2026-03-24)",
            "J": "Sprint 5 Implementation",
            "K": "Yash, Chaitanya",
            "L": "Chaitanya, Nayan",
            "M": "Complete",
            "N": "#22",
            "O": dt(2026, 3, 24),
            "P": dt(2026, 3, 24),
        },
    ),
    (
        10,
        {
            "A": "REQ-57",
            "B": "Patient & Records Manager",
            "C": "PatientController",
            "D": "getMyLabReportsList",
            "E": dt(2026, 3, 24),
            "F": dt(2026, 3, 24),
            "G": dt(2026, 3, 24),
            "H": "Added patient self-service access to their own lab reports with role-safe filtering",
            "I": "Commit 7718df7 (2026-03-24)",
            "J": "Sprint 5 Implementation",
            "K": "Yash, Chaitanya",
            "L": "Chaitanya, Nayan",
            "M": "Complete",
            "N": "#23",
            "O": dt(2026, 3, 24),
            "P": dt(2026, 3, 24),
        },
    ),
    (
        11,
        {
            "A": "REQ-56, REQ-57",
            "B": "Laboratory Manager",
            "C": "LabController",
            "D": "getLabRequestById",
            "E": dt(2026, 3, 24),
            "F": dt(2026, 3, 24),
            "G": dt(2026, 3, 24),
            "H": "Added single lab-request detail endpoint with role-based access checks",
            "I": "Commit f1ef67b (2026-03-24)",
            "J": "Sprint 5 Implementation",
            "K": "Chaitanya, Nayan",
            "L": "Chaitanya, Nayan",
            "M": "Complete",
            "N": "#24",
            "O": dt(2026, 3, 24),
            "P": dt(2026, 3, 24),
        },
    ),
]

for row, values in implementation_rows:
    set_row(ws, row, values)

# Shifted Sprint 4 rows should explicitly align with the official Sprint 4 window.
sprint4_rows = {
    12: {
        "E": dt(2026, 3, 8),
        "F": dt(2026, 3, 18),
        "G": dt(2026, 3, 18),
        "I": "Commit 51e610f (2026-03-18)",
        "J": "Sprint 4 Implementation",
    },
    13: {
        "E": dt(2026, 3, 8),
        "F": dt(2026, 3, 18),
        "G": dt(2026, 3, 18),
        "I": "Commit 5f625ec (2026-03-18)",
        "J": "Sprint 4 Implementation",
    },
    14: {
        "E": dt(2026, 3, 8),
        "F": dt(2026, 3, 18),
        "G": dt(2026, 3, 18),
        "I": "Commit fff4454 (2026-03-18)",
        "J": "Sprint 4 Implementation",
    },
    15: {
        "E": dt(2026, 3, 8),
        "F": dt(2026, 3, 18),
        "G": dt(2026, 3, 18),
        "I": "Commit 51d34ae (2026-03-18)",
        "J": "Sprint 4 Implementation",
    },
    16: {
        "E": dt(2026, 3, 8),
        "F": dt(2026, 3, 18),
        "G": dt(2026, 3, 18),
        "I": "Commit 17a6ac4 (2026-03-18)",
        "J": "Sprint 4 Implementation",
    },
    17: {
        "I": "Commit c8f72f8 (2026-02-27)",
    },
    18: {
        "H": "Added prescription creation and pharmacy dispensing",
        "I": "Commit d03bdc0 (2026-02-27)",
    },
    19: {
        "I": "Commit e200840 (2026-02-27)",
    },
    20: {
        "I": "Commit 8f634a0 (2026-02-27)",
    },
    21: {
        "I": "Commit 7f7b2d7 (2026-02-27)",
    },
    22: {
        "I": "Commit 01492a8 (2026-02-27)",
    },
    23: {
        "I": "Commit b7768e3 (2026-02-27)",
    },
    24: {
        "I": "Commit 40b6d1a (2026-02-27)",
    },
}

for row, values in sprint4_rows.items():
    set_row(ws, row, values)

ws["H18"] = "Added prescription creation and pharmacy dispensing"


# ---------------------------------------------------------------------------
# Testing
# ---------------------------------------------------------------------------
ws = wb["Testing"]
ws["F5"] = "Server running, doctor01 exists in DB and local LDAP is available"
ws["G5"] = 'POST /api/v1/auth/login with {ldapId: doctor01, password: doctor01pass}'


# ---------------------------------------------------------------------------
# Sprint Plans
# ---------------------------------------------------------------------------
ws = wb["Sprint Plans"]
set_row(
    ws,
    2,
    {
        "F": "Requirements baseline + UML + scope freeze (NFR-Architecture, NFR-Documentation)",
        "G": "SRS v1.0-v1.1, initial backlog, use-case/class/activity/component diagrams",
        "H": "SRS 1.0-1.1",
    },
)
set_row(
    ws,
    3,
    {
        "F": "LDAP auth + DB schema + patient profile (REQ-01, REQ-10, REQ-11, REQ-12, REQ-15, NFR-Security)",
        "G": "Auth flow, patient profile APIs, QR identification, schema design, SRS v1.2-v1.6",
        "H": "SRS 1.2-1.6",
    },
)
set_row(
    ws,
    4,
    {
        "F": "Visit lifecycle + doctor workflow (REQ-16, REQ-17, REQ-20, REQ-21, REQ-22, REQ-24, REQ-28)",
        "G": "Visit queue, consultation transitions, availability, check-in/out, attendance, SRS v2.0",
        "H": "SRS 2.0",
    },
)
set_row(
    ws,
    5,
    {
        "B": dt(2026, 2, 22),
        "C": dt(2026, 2, 28),
        "F": "Prescription + lab modules (REQ-26, REQ-44, REQ-45, REQ-46, REQ-48, REQ-54, REQ-55, REQ-56)",
        "G": "Prescription flow, dispensing flow, lab request/report flow, SRS 2.1 and testing refinement",
        "H": "SRS 2.0-2.1",
    },
)
set_row(
    ws,
    6,
    {
        "B": dt(2026, 3, 1),
        "C": dt(2026, 3, 18),
        "F": "Inventory + billing + appointments + QR check-in + documents (REQ-32, REQ-33, REQ-34, REQ-35, REQ-50, REQ-51, REQ-53)",
        "G": "Operational modules, smoke suite, integrated visit-to-billing flow, SRS 2.2 testing alignment",
        "H": "SRS 2.0-2.2",
    },
)
set_row(
    ws,
    7,
    {
        "F": "Admin + reports + notifications + lab access + LDAP integration (REQ-40, REQ-41, REQ-49, REQ-50, REQ-51, REQ-52, REQ-56, REQ-57, NFR-Integration, NFR-Documentation, NFR-Security)",
        "G": "Integrated backend build, E2E flow, API docs, availability notifications, reporting, local LDAP-backed auth",
        "H": "SRS 2.1-2.2 + testing-aligned backend closeout",
        "I": "Yash (integration), Chaitanya (testing/docs)",
        "J": "API + module integration demo with end-to-end flow",
    },
)
set_row(
    ws,
    8,
    {
        "F": "Frontend realization of existing backend FRs through role dashboards and route guards",
        "G": "Patient/doctor/reception/pharmacy/lab/admin UI screens mapped to current APIs",
        "H": "UI aligned to SRS actors and delivered backend requirements",
    },
)
set_row(
    ws,
    9,
    {
        "F": "Hardening + compliance verification (NFR-Performance, NFR-Security, NFR-Reliability, NFR-Deployment)",
        "G": "Security checks, RBAC validation, performance verification, deployment readiness",
        "H": "Final SRS compliance and release-readiness pass",
    },
)


# ---------------------------------------------------------------------------
# Sprint Updates
# ---------------------------------------------------------------------------
ws = wb["Sprint Updates"]
sprint_updates = {
    2: {
        "D": "Completed problem framing, initial SRS revisions, and the first UML baseline covering core PHC workflows and actors.",
        "E": "SRS 1.0, SRS 1.1, initial use-case/class/activity/component diagrams",
        "F": "Initial scope was too broad because ambulance-related ideas were still included.",
        "G": "Decided to narrow the project to a PHC-centered digital workflow system instead of a larger hospital-style platform.",
        "H": "Finalize architectural foundations, auth design, and database schema.",
        "I": 100,
        "J": "Yash / Nayan / Chaitanya",
        "K": dt(2026, 1, 24),
    },
    3: {
        "D": "Finalized the architecture baseline, schema direction, patient profile routes, JWT auth flow, and the SRS up to v2.0.",
        "E": "SRS 1.2-2.0, Prisma schema, auth middleware, patient profile endpoints, QR identification route",
        "F": "Database provisioning and full LDAP integration were not ready at the same time as the SRS freeze.",
        "G": "Adopted a layered route-controller-service-Prisma backend and kept LDAP fallback in dev until local bind infra was ready.",
        "H": "Implement the core visit lifecycle and doctor workflow.",
        "I": 100,
        "J": "Yash / Nayan / Chaitanya",
        "K": dt(2026, 2, 7),
    },
    4: {
        "D": "Implemented visit creation, vitals capture, queueing, claim/consult/complete transitions, and doctor availability/check-in/checkout flows.",
        "E": "Visit module, doctor/attendance module, route-level RBAC, Sprint 2 backend commits",
        "F": "Workflow correctness depended heavily on getting visit state transitions right before downstream modules.",
        "G": "Chose visit-centric linkage as the stable backbone for prescriptions, labs, and billing.",
        "H": "Extend the consultation flow into prescriptions and laboratory processing.",
        "I": 100,
        "J": "Yash / Chaitanya",
        "K": dt(2026, 2, 21),
    },
    5: {
        "D": "Delivered prescription creation/dispensing and lab request/report flows, then aligned SRS 2.1 and 2.2 with clearer NFR and testing sections.",
        "E": "Prescription module, lab module, SRS 2.1, SRS 2.2, Bruno collection growth",
        "F": "Role-based access for patient-visible lab data needed extra care after the core endpoints were implemented.",
        "G": "Used modular endpoint groups and Bruno-based verification before extending into operational modules.",
        "H": "Integrate operational modules: inventory, billing, appointments, QR check-in, and documents.",
        "I": 100,
        "J": "Chaitanya / Nayan / Yash",
        "K": dt(2026, 3, 7),
    },
    6: {
        "D": "Completed medicine inventory, billing with atomic stock deduction, appointments, QR check-in, external document handling, and formal smoke-test artifacts.",
        "E": "Billing module, medicine inventory, appointments, check-in, documents, smoke_test.sh, test_cases.csv, VHS recordings",
        "F": "Documentation synchronization became harder once testing artifacts and operational modules started landing together.",
        "G": "Used Prisma transactions for billing and check-in to keep multi-step operations consistent.",
        "H": "Finish admin and integration scope: user management, events, reports, notifications, real LDAP, and E2E validation.",
        "I": 100,
        "J": "Yash / Nayan / Chaitanya",
        "K": dt(2026, 3, 18),
    },
    7: {
        "D": "Completed Sprint 5 backend closeout: admin user management, PHC events, usage/attendance reports, patient lab access, notifications, E2E runner, API docs, and local LDAP-backed authentication.",
        "E": "Admin module, report endpoints, notifications, e2e-full-flow.js, API documentation, docker-compose LDAP setup",
        "F": "Most backend closeout work converged in a dense late-sprint batch, increasing pressure on docs and integration visibility.",
        "G": "Replaced the earlier LDAP stub behavior with a real local bind flow and added explicit end-to-end integration verification.",
        "H": "Begin Sprint 6 frontend dashboards while freezing backend scope except for blockers.",
        "I": 100,
        "J": "Yash / Chaitanya / Nayan",
        "K": dt(2026, 4, 3),
    },
    8: {
        "D": "No production frontend code merged yet. Sprint 6 scope is defined but implementation is intentionally pending while submissions and tracker updates are being completed.",
        "E": "Dashboard scope in README, Lab 9/10 report, role-wise frontend plan",
        "F": "Frontend work has not started in the repository; backend is ahead of UI delivery.",
        "G": "Hold backend scope steady and reuse the stable API surface for role dashboards rather than opening new backend features.",
        "H": "Patient, doctor, reception, pharmacy, lab, and admin dashboard implementation.",
        "I": 0,
        "J": "Chaitanya",
        "K": dt(2026, 4, 3),
    },
    9: {
        "D": "Not started yet.",
        "E": "Planned: security hardening, load/performance checks, RBAC testing, deployment plan",
        "F": "Pending completion of Sprint 6 UI work.",
        "G": "Sprint 7 remains a stabilization sprint, not a feature-expansion sprint.",
        "H": "Final testing, hardening, and deployment readiness.",
        "I": 0,
        "J": "All",
        "K": dt(2026, 4, 3),
    },
}

for row, values in sprint_updates.items():
    set_row(ws, row, values)


# ---------------------------------------------------------------------------
# Alpha Cards Progress - focused current-state updates supported by evidence
# ---------------------------------------------------------------------------
ws = wb["Alpha Cards Progress"]

alpha_updates = {
    34: {"D": 100, "E": 1, "G": "Lab 9/10 metrics now quantify response times, throughput, burn trends, and code size.", "H": TODAY},
    42: {"D": 100, "E": 1, "G": "Campus-scale backend completion and reused open-source tooling keep the solution practical for the project scope.", "H": TODAY},
    45: {"D": 75, "E": 0, "G": "The backend opportunity is substantially addressed through Sprint 5, but Sprint 6 frontend work is still pending.", "H": TODAY},
    56: {"D": 100, "E": 1, "G": "LDAP, integration, and documentation-drift risks are explicitly captured in the Lab 9/10 risk matrix.", "H": TODAY},
    57: {"D": 100, "E": 1, "G": "The backend architecture is demonstrable through live route modules, Prisma schema, and integrated flows.", "H": TODAY},
    58: {"D": 100, "E": 1, "G": "Performance has been measured through smoke timings, repeatability checks, and Lab 9/10 analysis.", "H": TODAY},
    60: {"D": 100, "E": 1, "G": "Critical interfaces including Prisma/Postgres, JWT auth, LDAP, and Bruno-tested APIs are demonstrated.", "H": TODAY},
    61: {"D": 100, "E": 1, "G": "Local LDAP, database migrations, and end-to-end HTTP flow demonstrate integration with the intended environment.", "H": TODAY},
    62: {"D": 75, "E": 0, "G": "Accepted as fit-for-purpose for backend scope; full product acceptance still depends on Sprint 6 and Sprint 7.", "H": TODAY},
    63: {"D": 75, "E": 0, "G": "The backend is operable through API flows, though the final user-facing frontend is not yet built.", "H": TODAY},
    64: {"D": 100, "E": 1, "G": "Functionality is covered by 48 documented test cases, expanded smoke coverage, and an end-to-end runner.", "H": TODAY},
    65: {"D": 100, "E": 1, "G": "Measured response times remain within stated thresholds for login, retrieval, billing, and QR check-in.", "H": TODAY},
    67: {"D": 75, "E": 0, "G": "Backend documentation is strong; final user/deployment documentation still belongs to later sprints.", "H": TODAY},
    68: {"D": 100, "E": 1, "G": "Sprint 0-5 backend release content is clear and documented in README, TESTING, and Lab 9/10 artifacts.", "H": TODAY},
    69: {"D": 100, "E": 1, "G": "Added value is clear through reduced paperwork, traceable visit linkage, and role-safe workflow automation.", "H": TODAY},
    108: {"D": 75, "E": 0, "G": "Enough requirements are addressed for backend acceptance, with frontend dashboards still pending.", "H": TODAY},
    109: {"D": 90, "E": 0, "G": "The implemented backend matches the major SRS workflows through Sprint 5.", "H": TODAY},
    110: {"D": 75, "E": 0, "G": "Value realization is visible in backend modules and test evidence, though the full user-facing product is incomplete.", "H": TODAY},
    111: {"D": 75, "E": 0, "G": "The system is worth making operational after frontend completion and Sprint 7 hardening.", "H": TODAY},
    125: {"D": 100, "E": 1, "G": "Risk exposure is now documented through the Lab 9/10 risk matrix and sprint-mitigation mapping.", "H": TODAY},
    137: {"D": 100, "E": 1, "G": "Sprint 0-5 backend tasks have been completed and tracked against the plan.", "H": TODAY},
    138: {"D": 75, "E": 0, "G": "Unplanned work stayed mostly contained to integration, docs sync, and LDAP hardening.", "H": TODAY},
    139: {"D": 75, "E": 0, "G": "Risks are identified and mitigated, but frontend lag and final hardening still remain.", "H": TODAY},
    140: {"D": 100, "E": 1, "G": "Schedule understanding has been revised to reflect backend-first delivery and deferred frontend scope.", "H": TODAY},
    141: {"D": 100, "E": 1, "G": "Progress is measured through git history, smoke data, E2E validation, and Lab 9/10 metrics.", "H": TODAY},
    142: {"D": 75, "E": 0, "G": "Some rework occurred when test coverage expanded and authentication moved from stub logic to local LDAP bind.", "H": TODAY},
    143: {"D": 100, "E": 1, "G": "Sprint 0-5 backend commitments were met; remaining work is intentionally Sprint 6/7 scope.", "H": TODAY},
    171: {"D": 100, "E": 1, "G": "Shared repo, Trello rhythm, testing assets, and report artifacts are accessible across the full team.", "H": TODAY},
    172: {"D": 100, "E": 1, "G": "The agreed way of working is available to the whole team through the repo, tracker, and documented process.", "H": TODAY},
    173: {"D": 100, "E": 1, "G": "The team has repeatedly inspected and adapted the process through SRS updates, testing additions, and Sprint 5 integration work.", "H": TODAY},
    174: {"D": 75, "E": 0, "G": "Progress is broadly predictable, though Sprint 5 still showed late convergence.", "H": TODAY},
    175: {"D": 75, "E": 0, "G": "Practices are consistently applied, but frontend work has not yet exercised the full delivery model.", "H": TODAY},
    176: {"D": 100, "E": 1, "G": "Git, Bruno, Prisma, Dockerized LDAP, and LaTeX/report tooling now support the team's workflow well.", "H": TODAY},
    177: {"D": 100, "E": 1, "G": "The way of working has been continually tuned through backend integration, testing expansion, and tracker updates.", "H": TODAY},
    204: {"D": 75, "E": 0, "G": "The team has met backend commitments consistently through Sprint 5, with frontend delivery still pending.", "H": TODAY},
    205: {"D": 100, "E": 1, "G": "Continuous improvement is visible in SRS revisions, smoke/e2e additions, and risk/metrics reporting.", "H": TODAY},
    206: {"D": 100, "E": 1, "G": "The team is operating with minimal supervision and self-managed sprint/accountability structure.", "H": TODAY},
    207: {"D": 75, "E": 0, "G": "The team handled scope reduction, LDAP hardening, and integration pressure effectively, with Sprint 6 still to prove UI execution.", "H": TODAY},
}

for row, values in alpha_updates.items():
    set_row(ws, row, values)


wb.save(PATH)
print(f"Updated {PATH}")
